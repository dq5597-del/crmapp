from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_DIR = Path(r"D:\Guanghui_AI_System\CRMAPP\outputs\01a0265e-5ae8-7211-9862-7ba1b66a7e1e")
SOURCE = OUTPUT_DIR / "光輝系統_產品批次匯入範本.xlsx"
TARGET = OUTPUT_DIR / "光輝系統_產品批次匯入範本_可直接貼圖片.xlsx"

wb = load_workbook(SOURCE)

image_columns = [
    ("AD", "主圖圖片", "可直接把一張圖片貼在同列儲存格內", 18),
    ("AE", "其他圖片", "可把多張圖片貼在同列儲存格範圍內", 24),
    ("AF", "產品介紹圖片", "可把多張介紹圖片貼在同列儲存格範圍內", 24),
]

header_fill = PatternFill("solid", fgColor="0F766E")
note_fill = PatternFill("solid", fgColor="E6F5F1")
white_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
note_font = Font(name="Arial", italic=True, color="52606D", size=9)
thin_gray = Side(style="thin", color="E5E7EB")

import_sheet = wb["產品匯入"]
import_sheet["B2"] = "必填"
for column, header, note, width in image_columns:
    import_sheet[f"{column}1"] = header
    import_sheet[f"{column}2"] = note
    import_sheet[f"{column}1"].fill = header_fill
    import_sheet[f"{column}1"].font = white_font
    import_sheet[f"{column}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    import_sheet[f"{column}2"].fill = note_fill
    import_sheet[f"{column}2"].font = note_font
    import_sheet[f"{column}2"].alignment = Alignment(vertical="center", wrap_text=True)
    import_sheet.column_dimensions[column].width = width
    for row in range(3, 203):
        cell = import_sheet[f"{column}{row}"]
        cell.border = Border(bottom=thin_gray, right=thin_gray)
        cell.alignment = Alignment(vertical="top")

example_sheet = wb["變體範例"]
for column, header, _, width in image_columns:
    example_sheet[f"{column}4"] = header
    example_sheet[f"{column}4"].fill = header_fill
    example_sheet[f"{column}4"].font = white_font
    example_sheet[f"{column}4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    example_sheet.column_dimensions[column].width = width
    for row in range(5, 9):
        example_sheet[f"{column}{row}"].border = Border(bottom=thin_gray, right=thin_gray)

field_sheet = wb["欄位說明"]
field_rows = [
    ["主圖圖片", "內嵌圖片", "選填", "直接把一張圖片插入或貼在同列儲存格內；會轉成 600×600 WebP", "特色／篩選／圖片"],
    ["其他圖片", "內嵌圖片", "選填", "可放多張；依圖片位置排序並轉成 600×600 WebP", "特色／篩選／圖片"],
    ["產品介紹圖片", "內嵌圖片", "選填", "可放多張；保留比例、轉成 WebP 並接在官網產品介紹文字後方", "特色／篩選／圖片"],
]
for row_values in field_rows:
    field_sheet.append(row_values)
for row in range(field_sheet.max_row - 2, field_sheet.max_row + 1):
    for column in range(1, 6):
        cell = field_sheet.cell(row, column)
        cell.font = Font(name="Arial", color="1F2937", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin_gray, right=thin_gray)
    field_sheet.cell(row, 3).fill = PatternFill("solid", fgColor="FFF7D6")
    field_sheet.cell(row, 5).fill = PatternFill("solid", fgColor="EEF2FF")

help_sheet = wb["使用說明"]
help_sheet.merge_cells("A25:F25")
help_sheet["A25"] = "直接在 Excel 貼圖片"
help_sheet["A25"].fill = PatternFill("solid", fgColor="2563EB")
help_sheet["A25"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
help_lines = [
    "主圖圖片：每個商品列放一張，系統會轉成 600×600 WebP。",
    "其他圖片：同一列可放多張，系統依圖片左上角的位置排序。",
    "產品介紹圖片：同一列可放多張，會保留比例、轉成 WebP，接在產品介紹文字後方。",
    "請使用 Excel 的插入圖片或直接貼上，並讓圖片左上角位於正確商品列；不要使用 IMAGE() 公式。",
    "變體商品的介紹圖請放在「系列主商品」那一列。Excel 內嵌圖片與外部拖曳照片可混用。",
]
for offset, text in enumerate(help_lines, start=26):
    help_sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=6)
    help_sheet.cell(offset, 1, text)
    help_sheet.cell(offset, 1).fill = PatternFill("solid", fgColor="F0FDF4")
    help_sheet.cell(offset, 1).font = Font(name="Arial", color="047857", size=10)
    help_sheet.cell(offset, 1).alignment = Alignment(vertical="center", wrap_text=True)
    help_sheet.row_dimensions[offset].height = 28

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                existing = cell.font
                cell.font = Font(
                    name="Arial",
                    size=existing.sz,
                    bold=existing.bold,
                    italic=existing.italic,
                    color=existing.color,
                    underline=existing.underline,
                )

wb.save(TARGET)

check = load_workbook(TARGET, data_only=False)
assert check["產品匯入"]["AD1"].value == "主圖圖片"
assert check["產品匯入"]["AE1"].value == "其他圖片"
assert check["產品匯入"]["AF1"].value == "產品介紹圖片"
assert check["產品匯入"]["B2"].value == "必填"
assert not any(
    isinstance(cell.value, str) and cell.value.startswith("=")
    for sheet in check.worksheets
    for row in sheet.iter_rows()
    for cell in row
)
print(TARGET)
